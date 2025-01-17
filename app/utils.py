# app/utils.py

import pandas as pd
import numpy as np
from tabulate import tabulate
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def load_transaction_types(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='Trn Model', nrows=0)
        all_columns = df.columns.tolist()
        transaction_types = [col for col in all_columns[8:] if not col.startswith('Unnamed')]
        transaction_types = sorted(set(transaction_types))
        logging.info(f"Transaction types loaded: {transaction_types}")
        return transaction_types
    except Exception as e:
        logging.error(f"Failed to load transaction types: {str(e)}")
        return []

def process_excel_file(file_path):
    try:
        logging.info(f"Loading data from {file_path}")
        df = pd.read_excel(file_path, sheet_name='Trn Model')
        logging.info("Processing Excel file")
        hierarchy = create_hierarchy(file_path)
        row_identifiers = generate_row_identifiers(df, hierarchy)
        df.index = row_identifiers
        df.index.name = 'Unique_ID'
        column_identifiers = generate_column_identifiers(df)
        df.columns = column_identifiers
        df = df.loc[:, ~df.columns.str.startswith('Unnamed_Unnamed')]
        logging.info("Finished processing Excel file.")
        return df
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        return pd.DataFrame()

def create_hierarchy(file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook['Trn Model']
        hierarchical_data = extract_hierarchical_data(sheet)
        filled_data = fill_hierarchical_data(hierarchical_data)
        hierarchy = []
        current_path = []
        for value, indent in filled_data:
            current_path = current_path[:indent] + [''] * (indent - len(current_path)) + [value]
            hierarchy.append('_'.join(filter(None, current_path)))
        logging.info(f"Hierarchy created with {len(hierarchy)} elements")
        return hierarchy
    except Exception as e:
        logging.error(f"Error creating hierarchy: {str(e)}")
        return []

def extract_hierarchical_data(sheet):
    return [(cell.value.strip() if cell.value else None, get_indentation_level(cell))
            for row in sheet.iter_rows(min_row=1, max_col=1) for cell in row]

def get_indentation_level(cell):
    return int(cell.alignment.indent) if cell.alignment and cell.alignment.indent else 0

def fill_hierarchical_data(hierarchical_data):
    filled_data = []
    last_value, last_indent = None, 0
    for value, indent in hierarchical_data:
        if value:
            last_value, last_indent = value, indent
        filled_data.append((last_value, last_indent))
    return filled_data

def generate_row_identifiers(df, hierarchy):
    try:
        def create_identifier(row, hier_value):
            variable = getattr(row, 'Variable', None)
            variable = variable if pd.notna(variable) and variable != '' else None
            return f"{hier_value}.{variable}" if variable else hier_value
        row_identifiers = [create_identifier(row, hier) for row, hier in zip(df.itertuples(index=False), hierarchy)]
        logging.info(f"Generated {len(row_identifiers)} row identifiers")
        return row_identifiers
    except Exception as e:
        logging.error(f"Error generating row identifiers: {str(e)}")
        return []

def generate_column_identifiers(df):
    try:
        column_identifiers = []
        last_valid_column_name = None
        for col in df.columns:
            column_name = str(col).strip() if pd.notna(col) and str(col).strip() else ""
            if "Unnamed" in str(col):
                column_name = last_valid_column_name if pd.notna(df.iloc[0][col]) and str(df.iloc[0][col]).strip() else f"Unnamed_{col}"
            else:
                last_valid_column_name = column_name
            first_row_value = str(df.iloc[0][col]).strip() if pd.notna(df.iloc[0][col]) and str(df.iloc[0][col]).strip() else ""
            column_identifier = f"{column_name}_{first_row_value}" if column_name and first_row_value else column_name or first_row_value
            column_identifiers.append(column_identifier if column_identifier else col)
        logging.info(f"Generated {len(column_identifiers)} column identifiers")
        return column_identifiers
    except Exception as e:
        logging.error(f"Error generating column identifiers: {str(e)}")
        return []

def present_transaction_data(df, transaction_type):
    try:
        processed_df = process_transaction_data(df, transaction_type)
        columns = ['Category', 'Variable'] + [col for col in processed_df.columns if col not in ['Unique_ID', 'Category', 'Variable']]
        processed_df = processed_df[columns].fillna('')
        table = tabulate(processed_df, headers='keys', tablefmt='psql', disable_numparse=True, showindex=False)
        table_lines = table.split('\n')
        header = table_lines[1]
        header_underline = table_lines[2]
        data_rows = [line for line in table_lines[3:] if not set(line).issubset(set('+-|'))]
        cleaned_table = '\n'.join([f"Transaction Type: {transaction_type}\n", header, header_underline] + data_rows)
        logging.info(f"Presented transaction data for type: {transaction_type}")
        return cleaned_table
    except Exception as e:
        logging.error(f"Error presenting transaction data: {str(e)}")
        return "Error presenting transaction data."

def process_transaction_data(df, transaction_type):
    try:
        logging.info(f"Initial DataFrame shape: {df.shape}")
        pattern = f'(?i){transaction_type.strip()}'
        logging.info(f"Using regex pattern: {pattern}")

        # Log the column names to inspect if there's any discrepancy
        logging.info(f"Column names: {df.columns.tolist()}")

        filtered_df = df.filter(regex=pattern)
        logging.info(f"Filtered DataFrame shape (after regex): {filtered_df.shape}")

        filtered_df = filtered_df.replace(r'^\s*$', np.nan, regex=True)
        logging.info(f"Filtered DataFrame shape (after replacing empty strings): {filtered_df.shape}")

        filtered_df = filtered_df.dropna(how='all')
        logging.info(f"Filtered DataFrame shape (after dropping all-NaN rows): {filtered_df.shape}")

        filtered_df = filtered_df.rename(columns=lambda x: x.split('_', 1)[-1])
        filtered_df = filtered_df.reset_index()

        filtered_df = filtered_df.assign(
            Category=lambda x: x['Unique_ID'].str.split('.').str[0],
            Variable=lambda x: x['Unique_ID'].str.split('.').str[1]
        ).dropna(subset=['Category', 'Variable'])

        logging.info(f"Processed transaction data for type: {transaction_type} with {filtered_df.shape[0]} rows")
        return filtered_df
    except Exception as e:
        logging.error(f"Error processing transaction data: {str(e)}")
        return pd.DataFrame()
